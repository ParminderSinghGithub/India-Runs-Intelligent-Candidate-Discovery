"""XLSX export for the official submission.

Produces outputs/submission.xlsx — content-identical to submission.csv but
in Excel format, validated to 100 rows.
"""

from __future__ import annotations

from pathlib import Path
from typing import Sequence

from src.models.submission_row import SubmissionRow


class XlsxExporter:
    """Export SubmissionRow objects to a formatted Excel workbook."""

    HEADER = ["candidate_id", "rank", "score", "reasoning"]
    EXPECTED_ROWS = 100

    def export(
        self,
        rows: Sequence[SubmissionRow],
        output_path: Path,
        *,
        validate: bool = True,
    ) -> Path:
        """Write rows to an Excel workbook and return the path.

        Args:
            rows: Submission rows (same set used for submission.csv).
            output_path: Destination .xlsx file path.
            validate: If True, raise ValueError when row count != 100.

        Returns:
            Resolved output path.

        Raises:
            ValueError: Row count mismatch when validate=True.
            ImportError: openpyxl not installed.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Alignment,
                Border,
                Font,
                PatternFill,
                Side,
            )
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for XLSX export. "
                "Run: pip install openpyxl"
            ) from exc

        if validate and len(rows) != self.EXPECTED_ROWS:
            raise ValueError(
                f"Expected exactly {self.EXPECTED_ROWS} rows for XLSX export; "
                f"got {len(rows)}."
            )

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submission"

        # ── header row ────────────────────────────────────────────────────
        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(bold=True, color="F1F5F9", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="334155")
        header_border = Border(
            left=thin, right=thin, top=thin, bottom=thin
        )

        for col_idx, col_name in enumerate(self.HEADER, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = header_border

        ws.row_dimensions[1].height = 24

        # ── data rows ──────────────────────────────────────────────────────
        data_font = Font(size=10, color="1E293B")
        alt_fill = PatternFill("solid", fgColor="F8FAFC")
        data_border = Border(
            left=Side(border_style="hair", color="CBD5E1"),
            right=Side(border_style="hair", color="CBD5E1"),
            bottom=Side(border_style="hair", color="E2E8F0"),
        )
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        center_alignment = Alignment(horizontal="center", vertical="top")

        for row_idx, row in enumerate(rows, start=2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()

            cells = [
                (1, row.candidate_id, center_alignment),
                (2, row.rank,         center_alignment),
                (3, round(float(row.score), 6), center_alignment),
                (4, row.reasoning,    wrap_alignment),
            ]
            for col_idx, value, alignment in cells:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.fill = fill
                cell.border = data_border
                cell.alignment = alignment

            ws.row_dimensions[row_idx].height = 52

        # ── column widths ──────────────────────────────────────────────────
        col_widths = [18, 8, 14, 100]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ── freeze header ──────────────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── auto-filter ────────────────────────────────────────────────────
        ws.auto_filter.ref = f"A1:D{len(rows) + 1}"

        wb.save(output_path)
        return output_path
